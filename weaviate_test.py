from constants import *
from vlite2.utils import chop_and_chunk
from sentence_transformers import SentenceTransformer
import os
import time
import timeit
from dotenv import load_dotenv
import weaviate
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

def ingest_one_w(questions):
    short_data_embeddings = SentenceTransformer(model_name).encode(SHORT_DATA).tolist()
    with questions.batch.dynamic() as batch:
        batch.add_object(
            properties={"test": "id0"},
            vector=short_data_embeddings
        )

def retrieve_w(questions, text):
    vector = SentenceTransformer(model_name).encode(text).tolist()
    response = questions.query.near_vector(
        near_vector=vector,
        limit=k
    )

def ingest_many_w(questions):
    long_data_chunked = chop_and_chunk(LONG_DATA, 512)
    long_data_embeddings = SentenceTransformer(model_name).encode(long_data_chunked).tolist()
    with questions.batch.dynamic() as batch:
        for i in range(len(long_data_chunked)):
            batch.add_object(
                properties={"test": f"id{i}"},
                vector=long_data_embeddings[i]
            )

if __name__ == "__main__":
    load_dotenv(dotenv_path='.env', verbose=True)
    client = weaviate.connect_to_wcs(
        cluster_url = os.getenv('W_URL'),
        auth_credentials=weaviate.auth.AuthApiKey(api_key=os.getenv('W_API_KEY')),
    )
    client.collections.delete("Question")
    client.close()

    start_time = time.time()
    
    ingest_one_w_time = timeit.timeit('ingest_one_w(questions)', 
                        setup='''
import weaviate
import weaviate.classes as wvc
import os
from dotenv import load_dotenv
from __main__ import ingest_one_w
load_dotenv(dotenv_path='.env', verbose=True)
client = weaviate.connect_to_wcs(
    cluster_url = os.getenv('W_URL'),
    auth_credentials=weaviate.auth.AuthApiKey(api_key=os.getenv('W_API_KEY')),
)
client.collections.delete("Question")
questions = client.collections.create(
    name="Question",
)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED ingest ONE W")
    
    retrieve_one_w_time = timeit.timeit('retrieve_w(questions, "hello")', 
                        setup='''
import weaviate
import weaviate.classes as wvc
import os
from dotenv import load_dotenv
from __main__ import ingest_one_w, retrieve_w
load_dotenv(dotenv_path='.env', verbose=True)
client = weaviate.connect_to_wcs(
    cluster_url = os.getenv('W_URL'),
    auth_credentials=weaviate.auth.AuthApiKey(api_key=os.getenv('W_API_KEY')),
)
client.collections.delete("Question")
questions = client.collections.create(
    name="Question",
)
ingest_one_w(questions)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED retrieve ONE W")

    ingest_many_w_time = timeit.timeit('ingest_many_w(questions)', 
                        setup='''
import weaviate
import weaviate.classes as wvc
import os
from dotenv import load_dotenv
from __main__ import ingest_many_w
load_dotenv(dotenv_path='.env', verbose=True)
client = weaviate.connect_to_wcs(
    cluster_url = os.getenv('W_URL'),
    auth_credentials=weaviate.auth.AuthApiKey(api_key=os.getenv('W_API_KEY')),
)
client.collections.delete("Question")
questions = client.collections.create(
    name="Question",
)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED ingest MANY W")

    retrieve_many_w_time = timeit.timeit('retrieve_w(questions, "civil law")', 
                        setup='''
import weaviate
import weaviate.classes as wvc
import os
from dotenv import load_dotenv
from __main__ import ingest_many_w, retrieve_w
load_dotenv(dotenv_path='.env', verbose=True)
client = weaviate.connect_to_wcs(
    cluster_url = os.getenv('W_URL'),
    auth_credentials=weaviate.auth.AuthApiKey(api_key=os.getenv('W_API_KEY')),
)
client.collections.delete("Question")
questions = client.collections.create(
    name="Question",
)
ingest_many_w(questions)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED retrieve MANY W")
    
    print(f"\n\nNumber of executions averaged over: {num_executions}")
    print(f"w ingest one: {ingest_one_w_time}")
    print(f"w retrieve one: {retrieve_one_w_time}")
    print(f"w ingest many: {ingest_many_w_time}")
    print(f"w retrieve many: {retrieve_many_w_time}\n\n")

    load_dotenv(dotenv_path='.env', verbose=True)
    client = weaviate.connect_to_wcs(
        cluster_url = os.getenv('W_URL'),
        auth_credentials=weaviate.auth.AuthApiKey(api_key=os.getenv('W_API_KEY')),
    )
    client.collections.delete("Question")
    client.close()
    
    end_time = time.time()
    print(f"Total time to run: {end_time - start_time} seconds")

    if not os.path.exists('benchmark.xlsx'):
        workbook = xlsxwriter.Workbook('benchmark.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()
        first_df = pd.DataFrame({'': ["ingest One", "retrieve One", "ingest Many", "retrieve Many"]})
        first = 1
    else:
        first = 0
    writer = pd.ExcelWriter('benchmark.xlsx', engine='openpyxl', mode = 'a', if_sheet_exists='overlay')
    workbook = load_workbook("benchmark.xlsx")
    writer.workbook = workbook
    df = pd.DataFrame({'Weaviate': [ingest_one_w_time, retrieve_one_w_time, ingest_many_w_time, retrieve_many_w_time]})
    writer.worksheets = {ws.title: ws for ws in workbook.worksheets}
    reader = pd.read_excel('benchmark.xlsx')
    if (first == 1):
        first_df.to_excel(writer, startcol=reader.shape[1], index = False)
    df.to_excel(writer, startcol=reader.shape[1] + first, index = False)
    writer.close()
    