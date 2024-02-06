from constants import *
import time
import glob
import os
import timeit
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

def ingest_one_v1(v):
    v.memorize(SHORT_DATA)

def retrieve_v1(v, text):
    v.remember(text)

def ingest_many_v1(v):
    v.memorize(text=LONG_DATA)

if __name__ == "__main__":
    start_time = time.time()
    
    ingest_one_v1_time = timeit.timeit('ingest_one_v1(v)', 
                        setup='''
from vlite import VLite
from __main__ import ingest_one_v1
v = VLite()
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED ingest ONE V1")

    retrieve_one_v1_time = timeit.timeit('retrieve_v1(v, "hello")', 
                        setup='''
from vlite import VLite
from __main__ import retrieve_v1
v = VLite()
v.memorize("Hello! My name is Ray. How are you?")
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED retrieve ONE V1")

    ingest_many_v1_time = timeit.timeit('ingest_many_v1(v)', 
                        setup='''
from vlite import VLite
from __main__ import ingest_many_v1
v = VLite()
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED ingest MANY V1")

    retrieve_many_v1_time = timeit.timeit('retrieve_v1(v, text)', 
                        setup='''
from vlite import VLite
from __main__ import retrieve_v1, ingest_many_v1
v = VLite()
ingest_many_v1(v)
text = "civil law"
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED retrieve MANY V1")

    print(f"\n\nNumber of executions averaged over: {num_executions}")
    print(f"v1 ingest one: {ingest_one_v1_time}")
    print(f"v1 retrieve one: {retrieve_one_v1_time}")
    print(f"v1 ingest many: {ingest_many_v1_time}")
    print(f"v1 retrieve many: {retrieve_many_v1_time}")

    files_to_delete = glob.glob('*.npz')
    for f in files_to_delete:
        os.remove(f)

    end_time = time.time()
    print(f"Total time to run: {end_time - start_time} seconds")

    if not os.path.exists('benchmark.xlsx'):
        workbook = xlsxwriter.Workbook('benchmark.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()
        first_df = pd.DataFrame({'': ["ingest One", "retrieve One", "ingest Many", "retrieve Many"]})
        first = 1
    else:
        first = 0
    writer = pd.ExcelWriter('benchmark.xlsx', engine='openpyxl', mode = 'a', if_sheet_exists='overlay')
    workbook = load_workbook("benchmark.xlsx")
    writer.workbook = workbook
    df = pd.DataFrame({'VLite1': [ingest_one_v1_time, retrieve_one_v1_time, ingest_many_v1_time, retrieve_many_v1_time]})
    writer.worksheets = {ws.title: ws for ws in workbook.worksheets}
    reader = pd.read_excel('benchmark.xlsx')
    if (first == 1):
        first_df.to_excel(writer, startcol=reader.shape[1], index = False)
    df.to_excel(writer, startcol=reader.shape[1] + first, index = False)
    writer.close()
