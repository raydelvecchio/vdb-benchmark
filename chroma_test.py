from constants import *
from chromadb.utils import embedding_functions
from vlite2.utils import chop_and_chunk
from vlite2.model import EmbeddingModel
import time
import timeit
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import os

embedding_model = EmbeddingModel()

def ingest_one_cdb(cdb):
    cdb.upsert(
        documents = [SHORT_DATA],
        embeddings = embedding_model.embed(texts=SHORT_DATA),
        ids = ["id0"],
    )

def retrieve_cdb(cdb, text):
    results = cdb.query(
        query_embeddings=embedding_model.embed(texts=[text]),
        n_results=k
    )

def ingest_many_cdb(cdb):
    long_data_chunked = chop_and_chunk(LONG_DATA, 512)
    cdb.upsert(
        documents = long_data_chunked,
        embeddings = embedding_model.embed(texts=long_data_chunked),
        ids = ["id"+str(i) for i in range(len(long_data_chunked))],
    )

if __name__ == "__main__":
    start_time = time.time()

    ingest_one_cdb_time = timeit.timeit('ingest_one_cdb(collection)', 
                        setup='''
import chromadb
from chromadb.utils import embedding_functions
from __main__ import ingest_one_cdb
client = chromadb.Client()
collection = client.create_collection(name="collection")
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED ingest ONE CDB")

    retrieve_one_cdb_time = timeit.timeit('retrieve_cdb(collection2, "hello")', 
                        setup='''
import chromadb
from chromadb.utils import embedding_functions
from __main__ import retrieve_cdb
client = chromadb.Client()
collection2 = client.create_collection(name="collection2")
collection2.add(
        documents = ["Hello! My name is Ray. How are you?"],
        ids = ["id1"],
    )
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED retrieve ONE CDB")

    ingest_many_cdb_time = timeit.timeit('ingest_many_cdb(collection3)', 
                        setup='''
import chromadb
from chromadb.utils import embedding_functions
from __main__ import ingest_many_cdb
client = chromadb.Client()
collection3 = client.create_collection(name="collection3")
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED ingest MANY CDB")

    retrieve_many_cdb_time = timeit.timeit('retrieve_cdb(collection4, "civil law")', 
                        setup='''
import chromadb
from chromadb.utils import embedding_functions
from __main__ import ingest_many_cdb, retrieve_cdb
client = chromadb.Client()
collection4 = client.create_collection(name="collection4")
ingest_many_cdb(collection4)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED retrieve MANY CDB")

    print(f"\n\nNumber of executions averaged over: {num_executions}")
    print(f"cdb ingest one: {ingest_one_cdb_time}")
    print(f"cdb retrieve one: {retrieve_one_cdb_time}")
    print(f"cdb ingest many: {ingest_many_cdb_time}")
    print(f"cdb retrieve many: {retrieve_many_cdb_time}")

    end_time = time.time()
    print(f"Total time to run: {end_time - start_time} seconds")

    if not os.path.exists('benchmark.xlsx'):
        workbook = xlsxwriter.Workbook('benchmark.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()
        first_df = pd.DataFrame({'': ["ingest One", "retrieve One", "ingest Many", "retrieve Many"]})
        first = 1
    else:
        first = 0
    writer = pd.ExcelWriter('benchmark.xlsx', engine='openpyxl', mode = 'a', if_sheet_exists='overlay')
    workbook = load_workbook("benchmark.xlsx")
    writer.workbook = workbook
    df = pd.DataFrame({'Chroma': [ingest_one_cdb_time, retrieve_one_cdb_time, ingest_many_cdb_time, retrieve_many_cdb_time]})
    writer.worksheets = {ws.title: ws for ws in workbook.worksheets}
    reader = pd.read_excel('benchmark.xlsx')
    if (first == 1):
        first_df.to_excel(writer, startcol=reader.shape[1], index = False)
    df.to_excel(writer, startcol=reader.shape[1] + first, index = False)
    writer.close()
