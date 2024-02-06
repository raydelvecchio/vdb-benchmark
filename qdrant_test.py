from constants import *
import time
import glob
import os
import timeit
from vlite2.utils import chop_and_chunk
from qdrant_client.http.models import PointStruct
from sentence_transformers import SentenceTransformer
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

def memorize_one_q(q):
    short_data_embeddings = SentenceTransformer(model_name).encode(SHORT_DATA).tolist()
    q.upsert(
        collection_name="test_collection",
        points=[PointStruct(id=0, vector=short_data_embeddings)]
    )

def remember_q(q, text):
    vectors = SentenceTransformer(model_name).encode(text).tolist()
    search_result = q.search(
        collection_name="test_collection", query_vector=vectors
    )

def memorize_many_q(q):
    long_data_chunked = chop_and_chunk(LONG_DATA, 512)
    long_data_embeddings = SentenceTransformer(model_name).encode(long_data_chunked).tolist()
    points = [PointStruct(id=i, vector=long_data_embeddings[i]) for i in range(len(long_data_embeddings))]
    q.upsert(
        collection_name="test_collection",
        points=points
    )

if __name__ == "__main__":
    start_time = time.time()
    
    memorize_one_q_time = timeit.timeit('memorize_one_q(qdrant)', 
                        setup='''
from qdrant_client import QdrantClient
from constants import *
from qdrant_client.http.models import Distance, VectorParams
from __main__ import memorize_one_q
qdrant = QdrantClient(":memory:")
qdrant.create_collection(
collection_name="test_collection",
vectors_config=VectorParams(size=dimension, distance=Distance.COSINE)
)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED MEMORIZE ONE Q")

    remember_one_q_time = timeit.timeit('remember_q(qdrant, "hello")', 
                        setup='''
from qdrant_client import QdrantClient
from constants import *
from qdrant_client.http.models import Distance, VectorParams
from __main__ import memorize_one_q, remember_q
qdrant = QdrantClient(":memory:")
qdrant.create_collection(
collection_name="test_collection",
vectors_config=VectorParams(size=dimension, distance=Distance.COSINE)
)
memorize_one_q(qdrant)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED REMEMBER ONE Q")

    memorize_many_q_time = timeit.timeit('memorize_many_q(qdrant)', 
                        setup='''
from qdrant_client import QdrantClient
from constants import *
from qdrant_client.http.models import Distance, VectorParams
from __main__ import memorize_many_q
qdrant = QdrantClient(":memory:")
qdrant.create_collection(
collection_name="test_collection",
vectors_config=VectorParams(size=dimension, distance=Distance.COSINE)
)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED MEMORIZE MANY Q")

    remember_many_q_time = timeit.timeit('remember_q(qdrant, "civil law")', 
                        setup='''
from qdrant_client import QdrantClient
from constants import *
from qdrant_client.http.models import Distance, VectorParams
from __main__ import memorize_many_q, remember_q
qdrant = QdrantClient(":memory:")
qdrant.create_collection(
collection_name="test_collection",
vectors_config=VectorParams(size=dimension, distance=Distance.COSINE)
)
memorize_many_q(qdrant)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED REMEMBER MANY Q")

    print(f"\n\nNumber of executions averaged over: {num_executions}")
    print(f"q memorize one: {memorize_one_q_time}")
    print(f"q remember one: {remember_one_q_time}")
    print(f"q memorize many: {memorize_many_q_time}")
    print(f"q remember many: {remember_many_q_time}")

    end_time = time.time()
    print(f"Total time to run: {end_time - start_time} seconds")

    if not os.path.exists('benchmark.xlsx'):
        workbook = xlsxwriter.Workbook('benchmark.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()
        first_df = pd.DataFrame({'': ["Memorize One", "Remember One", "Memorize Many", "Remember Many"]})
        first = 1
    else:
        first = 0
    writer = pd.ExcelWriter('benchmark.xlsx', engine='openpyxl', mode = 'a', if_sheet_exists='overlay')
    workbook = load_workbook("benchmark.xlsx")
    writer.workbook = workbook
    df = pd.DataFrame({'Qdrant': [memorize_one_q_time, remember_one_q_time, memorize_many_q_time, remember_many_q_time]})
    writer.worksheets = {ws.title: ws for ws in workbook.worksheets}
    reader = pd.read_excel('benchmark.xlsx')
    if (first == 1):
        first_df.to_excel(writer, startcol=reader.shape[1], index = False)
    df.to_excel(writer, startcol=reader.shape[1] + first, index = False)
    writer.close()