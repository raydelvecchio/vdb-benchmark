from constants import *
from pinecone import Pinecone
from vlite2.utils import chop_and_chunk
from sentence_transformers import SentenceTransformer
import time
import timeit
import os
from dotenv import load_dotenv
from pinecone import Pinecone
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

def memorize_one_pc(index):
    short_data_embeddings = SentenceTransformer(model_name).encode(SHORT_DATA).tolist()
    index.upsert(vectors=[
        {"id": "id0", "values": short_data_embeddings}
    ])

def remember_pc(index, text):
    vector = SentenceTransformer(model_name).encode(text).tolist()
    query_response = index.query(
        namespace="ns0",
        vector=vector,
        top_k=10
    )

def memorize_many_pc(index):
    long_data_chunked = chop_and_chunk(LONG_DATA, 512)
    long_data_embeddings = SentenceTransformer(model_name).encode(long_data_chunked).tolist()
    vectors = [{"id": "id"+str(i), "values": long_data_embeddings[i]} for i in range(len(long_data_chunked))]
    index.upsert(vectors=vectors)

if __name__ == "__main__":
    load_dotenv(dotenv_path='.env', verbose=True)
    pc = Pinecone(api_key=os.getenv('PC_API_KEY'))
    for i in pc.list_indexes():
        if i['name'] == "quickstart":
            pc.delete_index(i['name'])

    start_time = time.time()

    memorize_one_pc_time = timeit.timeit('memorize_one_pc(index)', 
                        setup='''
from pinecone import Pinecone, PodSpec
import os
from dotenv import load_dotenv
from constants import dimension
from __main__ import memorize_one_pc
load_dotenv(dotenv_path='.env', verbose=True)
pc = Pinecone(api_key=os.getenv('PC_API_KEY'))
for i in pc.list_indexes():
    if i['name'] == "quickstart":
        pc.delete_index(i['name'])
pc.create_index(name="quickstart", dimension=dimension, metric="cosine", spec=PodSpec(environment='gcp-starter'))
index = pc.Index("quickstart")
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED MEMORIZE ONE PC")

    remember_one_pc_time = timeit.timeit('remember_pc(index, "hello")', 
                        setup='''
from pinecone import Pinecone, PodSpec
from sentence_transformers import SentenceTransformer
import os
from constants import dimension
from dotenv import load_dotenv
from __main__ import memorize_one_pc, remember_pc
load_dotenv(dotenv_path='.env', verbose=True)
pc = Pinecone(api_key=os.getenv('PC_API_KEY'))
for i in pc.list_indexes():
    if i['name'] == "quickstart":
        pc.delete_index(i['name'])
pc.create_index(name="quickstart", dimension=dimension, metric="cosine", spec=PodSpec(environment='gcp-starter'))
index = pc.Index("quickstart")
memorize_one_pc(index)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED REMEMBER ONE PC")

    memorize_many_pc_time = timeit.timeit('memorize_many_pc(index)', 
                        setup='''
from pinecone import Pinecone, PodSpec
import os
from constants import dimension
from dotenv import load_dotenv
from __main__ import memorize_many_pc
load_dotenv(dotenv_path='.env', verbose=True)
pc = Pinecone(api_key=os.getenv('PC_API_KEY'))
for i in pc.list_indexes():
    if i['name'] == "quickstart":
        pc.delete_index(i['name'])
pc.create_index(name="quickstart", dimension=dimension, metric="cosine", spec=PodSpec(environment='gcp-starter'))
index = pc.Index("quickstart")
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED MEMORIZE MANY PC")

    remember_many_pc_time = timeit.timeit('remember_pc(index, "civil law")', 
                        setup='''
from pinecone import Pinecone, PodSpec
from sentence_transformers import SentenceTransformer
import os
from constants import dimension
from dotenv import load_dotenv
from __main__ import memorize_many_pc, remember_pc
load_dotenv(dotenv_path='.env', verbose=True)
pc = Pinecone(api_key=os.getenv('PC_API_KEY'))
for i in pc.list_indexes():
    if i['name'] == "quickstart":
        pc.delete_index(i['name'])
pc.create_index(name="quickstart", dimension=dimension, metric="cosine", spec=PodSpec(environment='gcp-starter'))
index = pc.Index("quickstart")
memorize_many_pc(index)
                        ''',
                        number=num_executions) / num_executions
    print("FINISHED REMEMBER MANY PC")

    print(f"\n\nNumber of executions averaged over: {num_executions}")
    print(f"pc memorize one: {memorize_one_pc_time}")
    print(f"pc remember one: {remember_one_pc_time}")
    print(f"pc memorize many: {memorize_many_pc_time}")
    print(f"pc remember many: {remember_many_pc_time}")

    load_dotenv(dotenv_path='.env', verbose=True)
    pc = Pinecone(api_key=os.getenv('PC_API_KEY'))
    for i in pc.list_indexes():
        if i['name'] == "quickstart":
            pc.delete_index(i['name'])

    end_time = time.time()
    print(f"Total time to run: {end_time - start_time} seconds")

    if not os.path.exists('benchmark.xlsx'):
        workbook = xlsxwriter.Workbook('benchmark.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()
        first_df = pd.DataFrame({'': ["Memorize One", "Remember One", "Memorize Many", "Remember Many"]})
        first = 1
    else:
        first = 0
    writer = pd.ExcelWriter('benchmark.xlsx', engine='openpyxl', mode = 'a', if_sheet_exists='overlay')
    workbook = load_workbook("benchmark.xlsx")
    writer.workbook = workbook
    df = pd.DataFrame({'Pinecone': [memorize_one_pc_time, remember_one_pc_time, memorize_many_pc_time, remember_many_pc_time]})
    writer.worksheets = {ws.title: ws for ws in workbook.worksheets}
    reader = pd.read_excel('benchmark.xlsx')
    if (first == 1):
        first_df.to_excel(writer, startcol=reader.shape[1], index = False)
    df.to_excel(writer, startcol=reader.shape[1] + first, index = False)
    writer.close()
